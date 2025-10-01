import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

# Hinweis an den Benutzer
st.info("Die rot gefärbten Zeilen müssen manuell eingetragen werden. Dispo!")

# Funktion zum Extrahieren der relevanten Daten für einen Bereich
def extract_work_data_for_range(df, start_value, end_value):
    relevant_words = [
        "ausgleich", "krank", "sonderurlaub", "urlaub",
        "berufsschule", "fahrschule", "homeoffice",
        "schulung", "dienstreise", "seminar", "fortbildung",
        "elternzeit", "kur", "reha", "kur und reha", "reha und kur"
    ]
    excluded_words = ["hoffahrer", "waschteam", "aushilfsfahrer"]
    result = []

    df.iloc[:, 1] = df.iloc[:, 1].astype(str).str.strip().str.lower()

    if start_value not in df.iloc[:, 1].values or end_value not in df.iloc[:, 1].values:
        st.error(f"Die Werte '{start_value}' oder '{end_value}' wurden in Spalte B nicht gefunden.")
        st.stop()

    start_index = df[df.iloc[:, 1] == start_value].index[0]
    end_index = df[df.iloc[:, 1] == end_value].index[0]

    for row_index in range(start_index, end_index + 1):
        lastname = str(df.iloc[row_index, 1]).strip().title()
        firstname = str(df.iloc[row_index, 2]).strip().title()

        if not lastname or not firstname or lastname == "None" or firstname == "None":
            continue

        activities_row = row_index + 1

        row = {
            "Nachname": lastname,
            "Vorname": firstname,
            "Sonntag": "",
            "Montag": "",
            "Dienstag": "",
            "Mittwoch": "",
            "Donnerstag": "",
            "Freitag": "",
            "Samstag": "",
        }

        for day, (col1, col2) in enumerate([(4, 5), (6, 7), (8, 9), (10, 11), (12, 13), (14, 15), (16, 17)]):
            activity1 = str(df.iloc[activities_row, col1]).strip()
            activity2 = str(df.iloc[activities_row, col2]).strip()
            activity = " ".join(filter(lambda x: x and x != "0", [activity1, activity2])).strip().lower()

            if (any(word in activity for word in relevant_words) and
                not any(excluded in activity for excluded in excluded_words)):
                weekday = ["Sonntag", "Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"][day]
                row[weekday] = activity.title()

        result.append(row)

    return pd.DataFrame(result)


# Funktion, um die Datumszeile zu erstellen
def create_header_with_dates(df):
    dates = [
        pd.to_datetime(df.iloc[1, 4]).strftime('%d.%m.%Y'),  # E2
        pd.to_datetime(df.iloc[1, 6]).strftime('%d.%m.%Y'),  # G2
        pd.to_datetime(df.iloc[1, 8]).strftime('%d.%m.%Y'),  # I2
        pd.to_datetime(df.iloc[1, 10]).strftime('%d.%m.%Y'), # K2
        pd.to_datetime(df.iloc[1, 12]).strftime('%d.%m.%Y'), # M2
        pd.to_datetime(df.iloc[1, 14]).strftime('%d.%m.%Y'), # O2
        pd.to_datetime(df.iloc[1, 16]).strftime('%d.%m.%Y'), # Q2
    ]
    return dates

def style_excel(ws, calendar_week, num_new_rows, total_rows):
    # Moderne Farbpalette
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dunkelblau für KW/Abteilung
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Mittelblau für Header
    data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Weiß
    data_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # Hellgrau
    
    # Rot-Töne für neue Mitarbeiter (müssen manuell eingetragen werden)
    new_row_fill_dark = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # Dunkelrot
    new_row_fill_light = PatternFill(start_color="F1948A", end_color="F1948A", fill_type="solid")  # Hellrot
    
    # Grün-Töne für letzte Zeilen
    last_row_fill_dark = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Dunkelgrün
    last_row_fill_light = PatternFill(start_color="82E0AA", end_color="82E0AA", fill_type="solid")  # Hellgrün
    
    # Rahmen
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    
    medium_border = Border(
        left=Side(style="medium", color="1F4E78"),
        right=Side(style="medium", color="1F4E78"),
        top=Side(style="medium", color="1F4E78"),
        bottom=Side(style="medium", color="1F4E78")
    )

    # KW-Eintrag oberhalb der Tabelle (Zeile 1)
    ws["A1"].value = f"Kalenderwoche: {calendar_week + 1}"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = title_fill
    ws["A1"].border = medium_border
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.row_dimensions[1].height = 30

    # Abteilung unterhalb der KW (Zeile 2)
    ws["A2"].value = "Abteilung: Fuhrpark - NFC"
    ws["A2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = title_fill
    ws["A2"].border = medium_border
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    ws.row_dimensions[2].height = 26

    # Header-Zeile (Zeile 3) - fett, zentriert und farbig
    for col in ws.iter_cols(min_row=3, max_row=3, min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = header_fill
            cell.border = medium_border
    ws.row_dimensions[3].height = 24

    # Datenzeilen formatieren (abwechselnd einfärben)
    for row in range(4, ws.max_row + 1):
        for cell in ws[row]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(size=10, color="2C3E50")
            cell.border = thin_border
            
            # Alternierende Farben für normale Zeilen
            if row % 2 == 0:
                cell.fill = data_fill_light
            else:
                cell.fill = data_fill_white
        
        ws.row_dimensions[row].height = 20

    # Formatierung für die ersten num_new_rows Datenzeilen (rot - manuell einzutragen)
    for row in range(4, 4 + num_new_rows):
        for cell in ws[row]:
            if (row - 4) % 2 == 0:  # Gerade Zeilen
                cell.fill = new_row_fill_dark
                cell.font = Font(size=10, color="FFFFFF", bold=True)
            else:  # Ungerade Zeilen
                cell.fill = new_row_fill_light
                cell.font = Font(size=10, color="2C3E50", bold=True)
            cell.border = thin_border

    # Formatierung für die letzten 7 Zeilen (grün)
    for row in range(ws.max_row - 6, ws.max_row + 1):
        for cell in ws[row]:
            relative_row = row - (ws.max_row - 6)
            if relative_row % 2 == 0:  # Gerade Zeilen
                cell.fill = last_row_fill_dark
                cell.font = Font(size=10, color="FFFFFF", bold=True)
            else:  # Ungerade Zeilen
                cell.fill = last_row_fill_light
                cell.font = Font(size=10, color="2C3E50", bold=True)
            cell.border = thin_border

    # Spaltenbreite anpassen mit Mindestbreiten
    column_min_widths = {
        1: 18,  # Nachname
        2: 16,  # Vorname
    }
    
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Berechne Breite mit Puffer
        calculated_width = max_length + 4
        
        # Verwende Mindestbreite falls definiert
        min_width = column_min_widths.get(col_idx, 15)
        adjusted_width = max(calculated_width, min_width)
        
        # Maximalbreite begrenzen
        adjusted_width = min(adjusted_width, 60)
        
        ws.column_dimensions[col_letter].width = adjusted_width

    # Erste drei Zeilen fixieren (KW, Abteilung, Header)
    ws.freeze_panes = "A4"


# Streamlit App
st.title("Wochenarbeitsbericht Fuhrpark")
uploaded_file = st.file_uploader("Lade eine Excel-Datei hoch", type=["xlsx"])

if uploaded_file:
    # Initialisiere Fortschrittsbalken
    progress_bar = st.progress(0)
    progress_status = st.empty()
    
    # Lade die Excel-Datei
    progress_status.text("Lade Excel-Datei...")
    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb["Druck Fahrer"]
    data = pd.DataFrame(sheet.values)
    progress_status.text("Excel-Daten geladen.")
    progress_bar.progress(30)

    # Erstelle 6 Zeilen für die Mitarbeiter oberhalb von "Adler"
    new_data = pd.DataFrame([{
        "Nachname": "Carstensen", "Vorname": "Martin", "Sonntag": "", "Montag": "", "Dienstag": "", 
        "Mittwoch": "", "Donnerstag": "", "Freitag": "", "Samstag": ""
    }, {
        "Nachname": "Richter", "Vorname": "Clemens", "Sonntag": "", "Montag": "", "Dienstag": "", 
        "Mittwoch": "", "Donnerstag": "", "Freitag": "", "Samstag": ""
    }, {
        "Nachname": "Lau", "Vorname": "Eike", "Sonntag": "", "Montag": "", "Dienstag": "", 
        "Mittwoch": "", "Donnerstag": "", "Freitag": "", "Samstag": ""
    }, {
        "Nachname": "Pham Manh", "Vorname": "Chris", "Sonntag": "", "Montag": "", "Dienstag": "", 
        "Mittwoch": "", "Donnerstag": "", "Freitag": "", "Samstag": ""
    }, {
        "Nachname": "Ohlenroth", "Vorname": "Nadja", "Sonntag": "", "Montag": "", "Dienstag": "", 
        "Mittwoch": "", "Donnerstag": "", "Freitag": "", "Samstag": ""
    }, {
        "Nachname": "Schulz", "Vorname": "Julian", "Sonntag": "", "Montag": "", "Dienstag": "", 
        "Mittwoch": "", "Donnerstag": "", "Freitag": "", "Samstag": ""
    }])

    # Extrahiere die relevanten Daten
    progress_status.text("Extrahiere Daten...")
    extracted_data_1 = extract_work_data_for_range(data, "adler", "steckel")
    progress_bar.progress(60)

    # Füge alle Daten zusammen
    extracted_data = pd.concat([new_data, extracted_data_1], ignore_index=True)

    # Kalenderwoche berechnen
    dates = create_header_with_dates(data)
    first_date = pd.to_datetime(dates[0], format='%d.%m.%Y')
    calendar_week = first_date.isocalendar()[1]

    # Flache Spaltenüberschriften erstellen
    columns = ["Nachname", "Vorname"] + [f"{weekday} ({date})" for weekday, date in zip(
        ["Sonntag", "Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"], dates
    )]
    extracted_data.columns = columns

    # Excel-Dateiname mit Kalenderwoche erstellen
    excel_filename = f"Fuhrpark_Meldung_KW: {calendar_week + 1}.xlsx"

    # Exportiere die Daten als Excel-Datei
    progress_status.text("Exportiere Excel-Datei...")
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        extracted_data.to_excel(writer, index=False, sheet_name="Wochenübersicht", startrow=2)
        ws = writer.sheets["Wochenübersicht"]
        style_excel(ws, calendar_week, len(new_data), len(extracted_data))
    excel_data = output.getvalue()

    # Fortschrittsanzeige abschließen
    progress_status.text("Fertig!")
    st.success("Verarbeitung abgeschlossen.")
    progress_bar.progress(100)

    # Download-Option
    st.download_button(
        label="Download als Excel",
        data=excel_data,
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
